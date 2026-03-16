from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from datetime import timedelta
import random
import logging
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken
from django.db.models import Count
from django.db.models.functions import TruncDate

from .models import User, PendingUser, EmailVerification, SubscriptionOnboarding
from lessons.models import Lesson
from support.models import SupportThread
from .serializers import (
    UserRegistrationSerializer,
    OTPVerificationSerializer,
    ResendOTPSerializer,
    LoginSerializer,
    PasswordResetSerializer,
    PasswordResetConfirmSerializer,
    SubscriptionOnboardingSerializer,
    SubscriptionPlacementTestSerializer,
)

logger = logging.getLogger(__name__)

def generate_otp():
    """Generate a 5-digit OTP"""
    return str(random.randint(10000, 99999))

def send_otp_email(email, name, otp):
    """Send OTP verification email"""
    subject = 'Verify your email - Daily Spanish'
    html_message = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <h2 style="color: #3B4BB1;">Welcome to Daily Spanish!</h2>
        <p>Hi {name},</p>
        <p>Thank you for signing up! Please verify your email address by entering the following code:</p>
        <div style="background-color: #f8f9fa; padding: 20px; text-align: center; margin: 20px 0; border-radius: 8px;">
            <h1 style="color: #F25A37; font-size: 32px; margin: 0; letter-spacing: 5px;">{otp}</h1>
        </div>
        <p>This code will expire in 10 minutes.</p>
        <p>If you didn't create an account, please ignore this email.</p>
        <p>¡Bienvenido a tu aventura en español!</p>
        <p>Best regards,<br>The Daily Spanish Team</p>
    </div>
    """
    
    try:
        send_mail(
            subject=subject,
            message=f'Your verification code is: {otp}',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=html_message,
            fail_silently=False,
        )
        return True
    except Exception as e:
        logger.error(f'Failed to send email to {email}: {str(e)}')
        return False

@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """Handle user registration and send OTP"""
    serializer = UserRegistrationSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(
            {'error': 'Validation failed', 'details': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    validated_data = serializer.validated_data
    email = validated_data['email']
    
    try:
        # Check if there's a pending verification for this email
        existing_verification = EmailVerification.objects.filter(
            email=email,
            verified=False,
            expires__gt=timezone.now()
        ).first()
        
        if existing_verification:
            return Response(
                {'error': 'A verification email has already been sent. Please check your inbox or wait before requesting a new one.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Hash password
        hashed_password = make_password(validated_data['password'])
        
        # Generate OTP
        otp = generate_otp()
        otp_expires = timezone.now() + timedelta(minutes=10)
        
        # Store verification data
        EmailVerification.objects.create(
            email=email,
            otp=otp,
            expires=otp_expires,
            verified=False
        )
        
        # Store pending user data
        PendingUser.objects.update_or_create(
            email=email,
            defaults={
                'name': validated_data['name'],
                'country': validated_data['country'],
                'native_language': validated_data['native_language'],
                'password': hashed_password,
            }
        )
        
        # Send OTP email
        if send_otp_email(email, validated_data['name'], otp):
            return Response(
                {
                    'message': 'Verification email sent successfully',
                    'email': email
                },
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                {'error': 'Failed to send verification email'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except Exception as e:
        logger.error(f'Registration error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    """Verify OTP and create user account"""
    serializer = OTPVerificationSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(
            {'error': 'Validation failed', 'details': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    email = serializer.validated_data['email']
    otp = serializer.validated_data['otp']
    
    try:
        # Find the verification record
        verification = EmailVerification.objects.filter(
            email=email,
            otp=otp,
            verified=False
        ).first()
        
        if not verification:
            return Response(
                {'error': 'Invalid or expired OTP'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if verification.is_expired():
            return Response(
                {'error': 'OTP has expired'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get pending user data
        pending_user = PendingUser.objects.filter(email=email).first()
        if not pending_user:
            return Response(
                {'error': 'No pending registration found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create the actual user
        user = User.objects.create(
            username=email,  # Use email as username
            email=email,
            name=pending_user.name,
            country=pending_user.country,
            native_language=pending_user.native_language,
            password=pending_user.password,  # Already hashed
            is_verified=True
        )
        
        # Mark verification as completed
        verification.verified = True
        verification.save()
        
        # Clean up pending user data
        pending_user.delete()
        
        # Auto-login: issue JWT tokens
        refresh = RefreshToken.for_user(user)
        access_token = refresh.access_token

        return Response(
            {
                'message': 'Email verified successfully',
                'access_token': str(access_token),
                'refresh_token': str(refresh),
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'country': user.country,
                    'native_language': user.native_language,
                    'nickname': user.nickname,
                    'gender': user.gender,
                    'age': user.age,
                    'profile_image': user.profile_image,
                    'companion_image': getattr(user, 'companion_image', None),
                    'referral_source': user.referral_source,
                    'legal_notice_accepted': user.legal_notice_accepted,
                }
            },
            status=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f'OTP verification error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([AllowAny])
def resend_otp(request):
    """Resend OTP for email verification"""
    serializer = ResendOTPSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(
            {'error': 'Validation failed', 'details': serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    email = serializer.validated_data['email']
    
    try:
        # Check if there's a pending user
        pending_user = PendingUser.objects.filter(email=email).first()
        if not pending_user:
            return Response(
                {'error': 'No pending registration found for this email'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if there's a recent verification attempt
        recent_verification = EmailVerification.objects.filter(
            email=email,
            created_at__gt=timezone.now() - timedelta(minutes=1)
        ).first()
        
        if recent_verification:
            return Response(
                {'error': 'Please wait at least 1 minute before requesting a new OTP'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate new OTP
        otp = generate_otp()
        otp_expires = timezone.now() + timedelta(minutes=10)
        
        # Create new verification record
        EmailVerification.objects.create(
            email=email,
            otp=otp,
            expires=otp_expires,
            verified=False
        )
        
        # Send OTP email
        if send_otp_email(email, pending_user.name, otp):
            return Response(
                {
                    'message': 'New verification code sent successfully',
                    'email': email
                },
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                {'error': 'Failed to send verification email'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except Exception as e:
        logger.error(f'Resend OTP error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    """Login user and return JWT tokens"""
    try:
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            password = serializer.validated_data['password']
            remember_me = serializer.validated_data.get('remember_me', False)

            user = User.objects.filter(email=email).first()
            if not user or not user.check_password(password):
                return Response({'error': 'Invalid email or password'}, status=status.HTTP_400_BAD_REQUEST)

            if getattr(user, 'is_blocked', False):
                return Response(
                    {'error': 'Your account has been blocked. Please contact support.', 'code': 'blocked'},
                    status=status.HTTP_403_FORBIDDEN
                )

            if not user.is_active:
                return Response(
                    {'error': 'Your account has been deactivated. Please contact support.', 'code': 'inactive'},
                    status=status.HTTP_403_FORBIDDEN
                )

            refresh = RefreshToken.for_user(user)
            access_token = refresh.access_token

            if remember_me:
                refresh.set_exp(lifetime=timedelta(days=30))

            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])

            return Response({
                'message': 'Login successful',
                'access_token': str(access_token),
                'refresh_token': str(refresh),
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'country': user.country,
                    'native_language': user.native_language,
                    'nickname': user.nickname,
                    'gender': user.gender,
                    'age': user.age,
                    'profile_image': user.profile_image,
                    'companion_image': getattr(user, 'companion_image', None),
                    'referral_source': user.referral_source,
                    'legal_notice_accepted': user.legal_notice_accepted,
                    'is_staff': user.is_staff,
                    'is_superuser': user.is_superuser,
                    'is_blocked': getattr(user, 'is_blocked', False),
                    'is_active': user.is_active,
                }
            }, status=status.HTTP_200_OK)
        else:
            return Response(
                {'error': 'Invalid data', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        logger.error(f'Login error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_user(request):
    """Logout user by blacklisting refresh token"""
    try:
        refresh_token = request.data.get('refresh_token')
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
            
        return Response(
            {'message': 'Logout successful'},
            status=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f'Logout error: {str(e)}')
        return Response(
            {'message': 'Logout successful'},  # Always return success for security
            status=status.HTTP_200_OK
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    """Send password reset email"""
    try:
        serializer = PasswordResetSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            
            # Check if user exists
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                # Don't reveal if email exists for security
                return Response(
                    {'message': 'If the email exists, a password reset link has been sent'},
                    status=status.HTTP_200_OK
                )
            
            # Generate OTP for password reset
            otp = generate_otp()
            otp_expires = timezone.now() + timedelta(minutes=10)
            
            # Delete any existing password reset requests
            EmailVerification.objects.filter(
                email=email,
                verified=False
            ).delete()
            
            # Create new password reset record
            EmailVerification.objects.create(
                email=email,
                otp=otp,
                expires=otp_expires,
                verified=False
            )
            
            # Send password reset email
            if send_password_reset_email(email, user.name, otp):
                return Response(
                    {'message': 'If the email exists, a password reset link has been sent'},
                    status=status.HTTP_200_OK
                )
            else:
                return Response(
                    {'error': 'Failed to send password reset email'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            return Response(
                {'error': 'Invalid data', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        logger.error(f'Forgot password error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password(request):
    """Reset password with OTP verification"""
    try:
        serializer = PasswordResetConfirmSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data['email']
            otp = serializer.validated_data['otp']
            new_password = serializer.validated_data['new_password']
            
            # Verify OTP
            try:
                verification = EmailVerification.objects.get(
                    email=email,
                    otp=otp,
                    verified=False,
                    expires__gt=timezone.now()
                )
            except EmailVerification.DoesNotExist:
                return Response(
                    {'error': 'Invalid or expired OTP'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get user and update password
            try:
                user = User.objects.get(email=email)
                user.password = make_password(new_password)
                user.save()
                
                # Mark verification as used
                verification.verified = True
                verification.save()
                
                return Response(
                    {'message': 'Password reset successful'},
                    status=status.HTTP_200_OK
                )
                
            except User.DoesNotExist:
                return Response(
                    {'error': 'User not found'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            return Response(
                {'error': 'Invalid data', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        logger.error(f'Reset password error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_profile(request):
    """Get current user profile"""
    try:
        user = request.user
        return Response({
            'id': user.id,
            'name': user.name,
            'email': user.email,
            'country': user.country,
            'native_language': user.native_language,
            'level': user.level,
            'nickname': user.nickname,
            'gender': user.gender,
            'age': user.age,
            'profile_image': user.profile_image,
            'companion_image': getattr(user, 'companion_image', None),
            'date_joined': user.date_joined,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'referral_source': user.referral_source,
            'legal_notice_accepted': user.legal_notice_accepted,
            'is_blocked': getattr(user, 'is_blocked', False),
            'is_active': user.is_active,
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f'Get profile error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_user_profile(request):
    """Update current user profile"""
    try:
        user = request.user
        data = request.data
        
        # Update allowed fields
        if 'name' in data:
            user.name = data['name']
        if 'country' in data:
            user.country = data['country']
        if 'native_language' in data:
            user.native_language = data['native_language']
        if 'level' in data:
            user.level = data['level']
        if 'nickname' in data:
            user.nickname = data['nickname']
        if 'gender' in data:
            user.gender = data['gender']
        if 'age' in data:
            user.age = int(data['age']) if data['age'] else None
        if 'profile_image' in data:
            user.profile_image = data['profile_image']
        if 'companion_image' in data:
            allowed = {c[0] for c in getattr(user, 'COMPANION_CHOICES', [])}
            val = data.get('companion_image')
            if val in allowed:
                user.companion_image = val
            else:
                return Response({'error': 'Invalid companion_image'}, status=status.HTTP_400_BAD_REQUEST)
        if 'referral_source' in data:
            user.referral_source = data['referral_source']
        if 'legal_notice_accepted' in data:
            accepted = bool(data['legal_notice_accepted'])
            user.legal_notice_accepted = accepted
            if accepted and not user.legal_notice_accepted_at:
                user.legal_notice_accepted_at = timezone.now()
            
        user.save()
        
        return Response({
            'message': 'Profile updated successfully',
            'user': {
                'id': user.id,
                'name': user.name,
                'email': user.email,
                'country': user.country,
                'native_language': user.native_language,
                'nickname': user.nickname,
                'gender': user.gender,
                'age': user.age,
                'profile_image': user.profile_image,
                'companion_image': getattr(user, 'companion_image', None),
                'referral_source': user.referral_source,
                'legal_notice_accepted': user.legal_notice_accepted,
            }
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f'Update profile error: {str(e)}')
        return Response(
            {'error': 'Internal server error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def subscription_status(request):
    user = request.user
    return Response(
        {
            'first_time_payment': not bool(getattr(user, 'has_used_subscription', False)),
            'has_used_subscription': bool(getattr(user, 'has_used_subscription', False)),
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_subscription_onboarding(request):
    user = request.user
    serializer = SubscriptionOnboardingSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({'error': 'Validation failed', 'details': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    defaults = serializer.validated_data
    onboarding, _ = SubscriptionOnboarding.objects.update_or_create(user=user, defaults=defaults)

    start_pref = onboarding.start_preference
    if start_pref == 'beginning':
        if not user.level:
            user.level = 'A1'
        user.has_used_subscription = True
        user.save()
        onboarding.suggested_level = user.level
        onboarding.save(update_fields=['suggested_level', 'updated_at'])
        return Response(
            {
                'message': 'Onboarding saved',
                'next': 'cart',
                'suggested_level': onboarding.suggested_level,
            },
            status=status.HTTP_200_OK,
        )

    return Response(
        {
            'message': 'Onboarding saved',
            'next': 'test',
        },
        status=status.HTTP_200_OK,
    )


def _suggest_level_from_score(score: int) -> str:
    if score <= 2:
        return 'A1'
    if score <= 4:
        return 'A2'
    if score <= 6:
        return 'B1'
    if score <= 8:
        return 'B2'
    return 'C1'


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_subscription_placement_test(request):
    user = request.user
    serializer = SubscriptionPlacementTestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({'error': 'Validation failed', 'details': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    plan_key = serializer.validated_data.get('plan_key')
    answers = serializer.validated_data['answers']

    correct = {
        'q1': 'soy',
        'q2': 'tengo',
        'q3': 'en',
        'q4': 'quiero',
        'q5': 'fui',
        'q6': 'estoy',
        'q7': 'hablamos',
        'q8': 'comprar',
        'q9': 'porque',
        'q10': 'mañana',
    }

    normalized = {str(a.get('id')): a.get('value') for a in answers}
    score = 0
    for qid, expected in correct.items():
        if normalized.get(qid) == expected:
            score += 1

    suggested_level = _suggest_level_from_score(score)

    onboarding, _ = SubscriptionOnboarding.objects.get_or_create(user=user)
    if plan_key:
        onboarding.plan_key = plan_key
    onboarding.test_answers = answers
    onboarding.test_score = score
    onboarding.suggested_level = suggested_level
    onboarding.save()

    user.level = suggested_level
    user.has_used_subscription = True
    user.save()

    return Response(
        {
            'message': 'Test submitted',
            'score': score,
            'total': len(correct),
            'suggested_level': suggested_level,
        },
        status=status.HTTP_200_OK,
    )


def send_password_reset_email(email, name, otp):
    """Send password reset email"""
    subject = 'Reset your password - Daily Spanish'
    html_message = f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
        <h2 style="color: #3B4BB1;">Password Reset - Daily Spanish</h2>
        <p>Hi {name},</p>
        <p>You requested to reset your password. Please use the following code to reset your password:</p>
        <div style="background-color: #f8f9fa; padding: 20px; text-align: center; margin: 20px 0; border-radius: 8px;">
            <h1 style="color: #F25A37; font-size: 32px; margin: 0; letter-spacing: 5px;">{otp}</h1>
        </div>
        <p>This code will expire in 10 minutes.</p>
        <p>If you didn't request a password reset, please ignore this email.</p>
        <p>Best regards,<br>The Daily Spanish Team</p>
    </div>
    """
    
    try:
        send_mail(
            subject=subject,
            message=f'Your password reset code is: {otp}',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            html_message=html_message,
            fail_silently=False,
        )
        return True
    except Exception as e:
        logger.error(f'Password reset email error: {str(e)}')
        return False

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def list_users(request):
    try:
        if not request.user.is_staff:
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        if request.method == 'POST':
            data = request.data or {}
            email = (data.get('email') or '').strip().lower()
            password = data.get('password') or ''
            name = (data.get('name') or '').strip()
            country = (data.get('country') or '').strip()
            native_language = (data.get('native_language') or '').strip()

            if not email or not password:
                return Response({'error': 'Email and password are required'}, status=status.HTTP_400_BAD_REQUEST)
            if len(password) < 6:
                return Response({'error': 'Password must be at least 6 characters'}, status=status.HTTP_400_BAD_REQUEST)
            if User.objects.filter(email=email).exists():
                return Response({'error': 'A user with this email already exists'}, status=status.HTTP_400_BAD_REQUEST)

            u = User(
                email=email,
                username=email,
                name=name or None,
                country=country or None,
                native_language=native_language or None,
                nickname=(data.get('nickname') or None),
                gender=(data.get('gender') or None),
                age=(data.get('age') or None),
                profile_image=(data.get('profile_image') or None),
                level=(data.get('level') or None),
                is_active=bool(data.get('is_active', True)),
                is_blocked=bool(data.get('is_blocked', False)),
                is_verified=True,
            )
            u.password = make_password(password)
            u.save()
            return Response({'message': 'User created', 'user_id': u.id}, status=status.HTTP_201_CREATED)

        qs = User.objects.filter(is_superuser=False).exclude(id=request.user.id)
        users = []
        for u in qs:
            users.append({
                'id': u.id,
                'name': u.name,
                'email': u.email,
                'country': u.country,
                'native_language': u.native_language,
                'level': u.level,
                'nickname': u.nickname,
                'gender': u.gender,
                'age': u.age,
                'profile_image': u.profile_image,
                'is_active': u.is_active,
                'is_blocked': getattr(u, 'is_blocked', False),
                'date_joined': u.date_joined,
                'last_login': u.last_login,
            })
        return Response({'users': users}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'List users error: {str(e)}')
        return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def user_detail(request, user_id: str):
    try:
        if not request.user.is_staff:
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        user = User.objects.filter(id=user_id, is_superuser=False).first()
        if not user:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        if user.id == request.user.id:
            return Response({'error': 'Cannot modify current admin'}, status=status.HTTP_400_BAD_REQUEST)

        if request.method == 'GET':
            return Response({
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'country': user.country,
                    'native_language': user.native_language,
                    'level': user.level,
                    'nickname': user.nickname,
                    'gender': user.gender,
                    'age': user.age,
                    'profile_image': user.profile_image,
                    'is_active': user.is_active,
                    'is_blocked': getattr(user, 'is_blocked', False),
                    'date_joined': user.date_joined,
                    'last_login': user.last_login,
                }
            }, status=status.HTTP_200_OK)

        if request.method == 'PATCH':
            data = request.data or {}
            if 'email' in data:
                next_email = (data.get('email') or '').strip().lower()
                if not next_email:
                    return Response({'error': 'Email cannot be empty'}, status=status.HTTP_400_BAD_REQUEST)
                exists = User.objects.filter(email=next_email).exclude(id=user.id).exists()
                if exists:
                    return Response({'error': 'A user with this email already exists'}, status=status.HTTP_400_BAD_REQUEST)
                user.email = next_email
                user.username = next_email

            for field in ['name', 'country', 'native_language', 'nickname', 'gender', 'profile_image', 'level']:
                if field in data:
                    val = data.get(field)
                    setattr(user, field, val if val not in ['', None] else None)

            if 'age' in data:
                val = data.get('age')
                try:
                    user.age = int(val) if val not in ['', None] else None
                except Exception:
                    return Response({'error': 'Age must be a number'}, status=status.HTTP_400_BAD_REQUEST)

            if 'is_active' in data:
                user.is_active = bool(data.get('is_active'))
            if 'is_blocked' in data:
                user.is_blocked = bool(data.get('is_blocked'))

            if 'password' in data and data.get('password'):
                pwd = data.get('password')
                if len(pwd) < 6:
                    return Response({'error': 'Password must be at least 6 characters'}, status=status.HTTP_400_BAD_REQUEST)
                user.password = make_password(pwd)

            user.save()
            return Response({'message': 'User updated'}, status=status.HTTP_200_OK)

        if request.method == 'DELETE':
            user.delete()
            return Response({'message': 'User deleted'}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'User detail error: {str(e)}')
        return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    try:
        if not request.user.is_staff:
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        total_users = User.objects.filter(is_superuser=False).count()
        active_users = User.objects.filter(is_superuser=False, is_active=True, is_blocked=False).count()

        # Payments not implemented yet; return 0
        active_payments = 0
        overdue_payments = 0

        total_lessons = Lesson.objects.count()
        open_threads = SupportThread.objects.filter(status='open').count()
        resolved_threads = SupportThread.objects.filter(status='resolved').count()

        return Response({
            'total_users': total_users,
            'active_users': active_users,
            'active_payments': active_payments,
            'overdue_payments': overdue_payments,
            'total_lessons': total_lessons,
            'support_open': open_threads,
            'support_resolved': resolved_threads,
        }, status=status.HTTP_200_OK)
    except Exception:
        return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_analytics(request):
    try:
        if not request.user.is_staff:
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        days_raw = request.query_params.get('days') or '30'
        try:
            days = int(days_raw)
        except Exception:
            days = 30
        if days < 7:
            days = 7
        if days > 365:
            days = 365

        today = timezone.now().date()
        start_date = today - timedelta(days=days - 1)

        users_qs = User.objects.filter(is_superuser=False)

        total_users = users_qs.count()
        active_users = users_qs.filter(is_active=True, is_blocked=False).count()
        blocked_users = users_qs.filter(is_blocked=True).count()
        verified_users = users_qs.filter(is_verified=True).count()
        subscribed_users = users_qs.filter(has_used_subscription=True).count()

        countries = (
            users_qs.values('country')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        languages = (
            users_qs.values('native_language')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        levels = (
            users_qs.values('level')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        def normalize_bucket(row, field):
            raw = row.get(field)
            key = (raw or '').strip()
            label = key or 'Unknown'
            return {'key': key or 'unknown', 'label': label, 'count': int(row.get('count') or 0)}

        def top_n_with_other(rows, field, n=8):
            items = [normalize_bucket(r, field) for r in rows]
            top = items[:n]
            rest = items[n:]
            other_count = sum(i['count'] for i in rest)
            if other_count:
                top.append({'key': 'other', 'label': 'Other', 'count': other_count})
            return top

        by_country = top_n_with_other(countries, 'country', n=8)
        by_language = top_n_with_other(languages, 'native_language', n=8)
        by_level = top_n_with_other(levels, 'level', n=6)

        countries_count = sum(1 for r in countries if (r.get('country') or '').strip())
        languages_count = sum(1 for r in languages if (r.get('native_language') or '').strip())

        signups_qs = (
            users_qs.filter(date_joined__date__gte=start_date, date_joined__date__lte=today)
            .annotate(d=TruncDate('date_joined'))
            .values('d')
            .annotate(count=Count('id'))
            .order_by('d')
        )
        signup_map = {row['d']: int(row['count'] or 0) for row in signups_qs}
        signups_series = []
        for i in range(days):
            d = start_date + timedelta(days=i)
            signups_series.append({'date': d.isoformat(), 'count': signup_map.get(d, 0)})

        lessons_qs = Lesson.objects.all()
        lessons_total = lessons_qs.count()
        lessons_by_block = {b: lessons_qs.filter(block=b).count() for b in ['A1', 'A2', 'B1', 'B2', 'C1']}
        lessons_with_video_file = lessons_qs.exclude(video_file='').exclude(video_file__isnull=True).count()
        lessons_with_video_url = lessons_qs.exclude(video_url='').exclude(video_url__isnull=True).count()
        lessons_with_lesson_pdf = lessons_qs.exclude(lesson_pdf='').exclude(lesson_pdf__isnull=True).count()
        lessons_with_keys_pdf = lessons_qs.exclude(keys_pdf='').exclude(keys_pdf__isnull=True).count()

        threads_qs = SupportThread.objects.all()
        support_open = threads_qs.filter(status='open').count()
        support_resolved = threads_qs.filter(status='resolved').count()
        support_closed = threads_qs.filter(status='closed').count()

        recent_users = []
        for u in users_qs.order_by('-date_joined')[:12]:
            recent_users.append({
                'id': u.id,
                'email': u.email,
                'name': u.name,
                'country': u.country,
                'native_language': u.native_language,
                'level': u.level,
                'is_active': u.is_active,
                'is_blocked': getattr(u, 'is_blocked', False),
                'date_joined': u.date_joined,
                'last_login': u.last_login,
            })

        return Response(
            {
                'generated_at': timezone.now(),
                'window_days': days,
                'users': {
                    'total': total_users,
                    'active': active_users,
                    'blocked': blocked_users,
                    'verified': verified_users,
                    'subscribed': subscribed_users,
                    'countries_count': countries_count,
                    'languages_count': languages_count,
                    'by_country': by_country,
                    'by_language': by_language,
                    'by_level': by_level,
                    'signups': signups_series,
                },
                'lessons': {
                    'total': lessons_total,
                    'by_block': lessons_by_block,
                    'with_video_file': lessons_with_video_file,
                    'with_video_url': lessons_with_video_url,
                    'with_lesson_pdf': lessons_with_lesson_pdf,
                    'with_keys_pdf': lessons_with_keys_pdf,
                },
                'support': {
                    'total': support_open + support_resolved + support_closed,
                    'open': support_open,
                    'resolved': support_resolved,
                    'closed': support_closed,
                },
                'recent_users': recent_users,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        logger.error(f'Reports analytics error: {str(e)}')
        return Response({'error': 'Internal server error'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _payments_plan_details(plan_key: str | None):
    key = (plan_key or '').strip().lower()
    if key == 'yearly':
        return {'plan_key': 'yearly', 'plan_label': 'Annual', 'amount_cents': 19700, 'interval_days': 365}
    return {'plan_key': 'monthly', 'plan_label': 'Monthly', 'amount_cents': 2500, 'interval_days': 30}


def _payments_subscription_start(user: User, onboarding: SubscriptionOnboarding | None):
    if onboarding and onboarding.created_at:
        return onboarding.created_at
    if getattr(user, 'created_at', None):
        return user.created_at
    return user.date_joined


def _payments_effective_due_at(user: User, start_at, plan: dict):
    override = getattr(user, 'payment_due_override', None)
    if override is not None:
        return override
    if not start_at:
        return None
    return start_at + timedelta(days=plan['interval_days'])


def _pdf_escape(value: str) -> str:
    return (value or '').replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _build_simple_pdf(lines: list[str]) -> bytes:
    text_lines = [l if isinstance(l, str) else str(l) for l in (lines or [])]
    y_start = 770
    line_gap = 16
    commands = ['BT', '/F1 12 Tf', f'50 {y_start} Td']
    for idx, line in enumerate(text_lines):
        if idx > 0:
            commands.append(f'0 {-line_gap} Td')
        commands.append(f'({_pdf_escape(line)}) Tj')
    commands.append('ET')
    stream = ('\n'.join(commands) + '\n').encode('latin-1', errors='ignore')

    def obj(n: int, body: bytes) -> bytes:
        return f'{n} 0 obj\n'.encode('ascii') + body + b'\nendobj\n'

    objects: list[bytes] = []
    objects.append(obj(1, b'<< /Type /Catalog /Pages 2 0 R >>'))
    objects.append(obj(2, b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>'))
    objects.append(
        obj(
            3,
            b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>',
        )
    )
    objects.append(obj(4, b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>'))
    objects.append(
        obj(
            5,
            b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'endstream',
        )
    )

    header = b'%PDF-1.4\n%\xe2\xe3\xcf\xd3\n'
    body = bytearray()
    body.extend(header)

    xref_offsets = [0]
    for o in objects:
        xref_offsets.append(len(body))
        body.extend(o)

    xref_start = len(body)
    body.extend(b'xref\n')
    body.extend(f'0 {len(xref_offsets)}\n'.encode('ascii'))
    body.extend(b'0000000000 65535 f \n')
    for off in xref_offsets[1:]:
        body.extend(f'{off:010d} 00000 n \n'.encode('ascii'))

    body.extend(b'trailer\n')
    body.extend(f'<< /Size {len(xref_offsets)} /Root 1 0 R >>\n'.encode('ascii'))
    body.extend(b'startxref\n')
    body.extend(f'{xref_start}\n'.encode('ascii'))
    body.extend(b'%%EOF\n')
    return bytes(body)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_payments_overview(request):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    now = timezone.now()
    users_qs = User.objects.filter(is_superuser=False).select_related('subscription_onboarding').order_by('-date_joined')

    users = []
    active_subscriptions = []
    overdue_subscriptions = []

    est_total_revenue_cents = 0
    est_mrr_cents = 0

    for u in users_qs:
        onboarding = getattr(u, 'subscription_onboarding', None)
        plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
        has_sub = bool(getattr(u, 'has_used_subscription', False))
        start_at = _payments_subscription_start(u, onboarding) if has_sub else None
        next_due = _payments_effective_due_at(u, start_at, plan)
        is_overdue = bool(has_sub and next_due and next_due < now)

        status_key = 'none'
        if has_sub and is_overdue:
            status_key = 'overdue'
        elif has_sub:
            status_key = 'active'

        user_row = {
            'id': u.id,
            'name': u.name,
            'email': u.email,
            'plan_key': plan['plan_key'] if has_sub else None,
            'plan_label': plan['plan_label'] if has_sub else None,
            'amount_cents': plan['amount_cents'] if has_sub else 0,
            'currency': 'USD',
            'status': status_key,
            'started_at': start_at.isoformat() if start_at else None,
            'next_due': next_due.isoformat() if next_due else None,
            'date_joined': u.date_joined.isoformat() if u.date_joined else None,
        }
        users.append(user_row)

        if has_sub:
            est_total_revenue_cents += int(plan['amount_cents'])
            est_mrr_cents += int(plan['amount_cents']) if plan['plan_key'] == 'monthly' else int(round(plan['amount_cents'] / 12))

            sub_row = {
                'user_id': u.id,
                'user_name': u.name,
                'user_email': u.email,
                'plan_label': plan['plan_label'],
                'amount_cents': plan['amount_cents'],
                'currency': 'USD',
                'start_date': start_at.date().isoformat() if start_at else None,
                'next_payment': next_due.date().isoformat() if next_due else None,
                'status': 'Overdue' if is_overdue else 'Active',
            }
            if is_overdue:
                days_past = (now.date() - next_due.date()).days if next_due else 0
                overdue_subscriptions.append(
                    {
                        'user_id': u.id,
                        'user_name': u.name,
                        'email': u.email,
                        'plan_label': plan['plan_label'],
                        'amount_cents': plan['amount_cents'],
                        'currency': 'USD',
                        'due_date': next_due.date().isoformat() if next_due else None,
                        'days_past_due': days_past,
                    }
                )
            else:
                active_subscriptions.append(sub_row)

    plan_counts = {'monthly': 0, 'yearly': 0}
    for r in users:
        if r['plan_key'] == 'monthly':
            plan_counts['monthly'] += 1
        elif r['plan_key'] == 'yearly':
            plan_counts['yearly'] += 1

    return Response(
        {
            'generated_at': now.isoformat(),
            'stats': {
                'estimated_total_revenue_cents': est_total_revenue_cents,
                'estimated_mrr_cents': est_mrr_cents,
                'active_subscriptions': len(active_subscriptions),
                'overdue_payments': len(overdue_subscriptions),
                'subscribers_total': plan_counts['monthly'] + plan_counts['yearly'],
            },
            'plan_breakdown': plan_counts,
            'users': users,
            'subscriptions': active_subscriptions,
            'overdue': overdue_subscriptions,
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_send_payment_reminders(request):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    from notifications.models import Notification, NotificationRecipient

    now = timezone.now()
    users_qs = User.objects.filter(is_superuser=False, has_used_subscription=True).select_related('subscription_onboarding')
    overdue_ids: list[str] = []

    for u in users_qs:
        onboarding = getattr(u, 'subscription_onboarding', None)
        plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
        start_at = _payments_subscription_start(u, onboarding)
        next_due = _payments_effective_due_at(u, start_at, plan)
        if next_due and next_due < now:
            overdue_ids.append(u.id)

    title = (request.data or {}).get('title') if isinstance(request.data, dict) else None
    message = (request.data or {}).get('message') if isinstance(request.data, dict) else None
    title = (title or 'Payment overdue').strip()
    message = (message or 'Your subscription payment is overdue. Please update your payment method to continue uninterrupted access.').strip()

    notif = Notification.objects.create(type='alert', title=title, message=message, audience_filters={'mode': 'overdue'}, created_by=request.user)

    existing = set(
        NotificationRecipient.objects.filter(notification=notif, user_id__in=overdue_ids).values_list('user_id', flat=True)
    )
    to_create = []
    for uid in overdue_ids:
        if uid in existing:
            continue
        to_create.append(NotificationRecipient(notification=notif, user_id=uid, delivered_at=now))
    NotificationRecipient.objects.bulk_create(to_create, batch_size=1000)
    notif.sent_at = now
    notif.save(update_fields=['sent_at'])

    return Response({'sent_to': len(to_create), 'overdue_users': len(overdue_ids)}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_invoices(request):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    import io
    import zipfile
    from django.http import HttpResponse

    now = timezone.now()
    users_qs = User.objects.filter(is_superuser=False).select_related('subscription_onboarding').order_by('email')

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for u in users_qs:
            onboarding = getattr(u, 'subscription_onboarding', None)
            has_sub = bool(getattr(u, 'has_used_subscription', False))
            plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
            start_at = _payments_subscription_start(u, onboarding) if has_sub else None
            next_due = _payments_effective_due_at(u, start_at, plan)
            is_overdue = bool(has_sub and next_due and next_due < now)

            issued = now.date().isoformat()
            invoice_number = f'INV-{now.strftime("%Y%m%d")}-{u.id[-6:]}'
            amount_cents = plan['amount_cents'] if has_sub else 0

            lines = [
                'Daily Spanish',
                'Invoice',
                '',
                f'Invoice #: {invoice_number}',
                f'Issued on: {issued}',
                '',
                f'Bill to: {(u.name or "").strip() or "User"}',
                f'Email: {u.email}',
                '',
                f'Plan: {plan["plan_label"] if has_sub else "N/A"}',
                f'Amount: ${amount_cents / 100:.2f} USD',
                f'Status: {"Overdue" if is_overdue else "Paid" if has_sub else "No subscription record"}',
                f'Next due: {next_due.date().isoformat() if next_due else "N/A"}',
            ]
            pdf_bytes = _build_simple_pdf(lines)
            safe_email = (u.email or u.id).replace('@', '_at_').replace('/', '_')
            zf.writestr(f'{safe_email}_{invoice_number}.pdf', pdf_bytes)

    zip_bytes = buf.getvalue()
    resp = HttpResponse(zip_bytes, content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="invoices_{now.strftime("%Y%m%d_%H%M%S")}.zip"'
    return resp


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_export_financial_report(request):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    now = timezone.now()
    users_qs = User.objects.filter(is_superuser=False).select_related('subscription_onboarding').order_by('email')

    subscribers = []
    overdue = []
    plan_counts = {'monthly': 0, 'yearly': 0}
    est_total_revenue_cents = 0
    est_mrr_cents = 0

    for u in users_qs:
        onboarding = getattr(u, 'subscription_onboarding', None)
        has_sub = bool(getattr(u, 'has_used_subscription', False))
        plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
        if has_sub:
            plan_counts[plan['plan_key']] += 1
            est_total_revenue_cents += int(plan['amount_cents'])
            est_mrr_cents += int(plan['amount_cents']) if plan['plan_key'] == 'monthly' else int(round(plan['amount_cents'] / 12))

        start_at = _payments_subscription_start(u, onboarding) if has_sub else None
        next_due = _payments_effective_due_at(u, start_at, plan)
        is_overdue = bool(has_sub and next_due and next_due < now)

        row = {
            'id': u.id,
            'name': u.name,
            'email': u.email,
            'plan': plan['plan_label'] if has_sub else '',
            'amount_usd': float(plan['amount_cents'] / 100) if has_sub else 0.0,
            'started_at': start_at.date().isoformat() if start_at else '',
            'next_due': next_due.date().isoformat() if next_due else '',
            'status': 'Overdue' if is_overdue else 'Active' if has_sub else 'None',
        }
        subscribers.append(row)
        if is_overdue:
            overdue.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='3B4BB1')
    ws['A1'] = 'Daily Spanish — Financial Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Generated at (UTC): {now.strftime("%Y-%m-%d %H:%M:%S")}'

    ws['A4'] = 'Estimated total revenue (USD)'
    ws['B4'] = float(est_total_revenue_cents / 100)
    ws['A5'] = 'Estimated MRR (USD)'
    ws['B5'] = float(est_mrr_cents / 100)
    ws['A6'] = 'Monthly subscribers'
    ws['B6'] = plan_counts['monthly']
    ws['A7'] = 'Annual subscribers'
    ws['B7'] = plan_counts['yearly']
    ws['A8'] = 'Overdue subscribers'
    ws['B8'] = len(overdue)

    ws2 = wb.create_sheet('Subscriptions')
    ws3 = wb.create_sheet('Overdue')

    def write_table(sheet, rows):
        cols = ['id', 'name', 'email', 'plan', 'amount_usd', 'started_at', 'next_due', 'status']
        for c_idx, c in enumerate(cols, start=1):
            cell = sheet.cell(row=1, column=c_idx, value=c.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, c in enumerate(cols, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=row.get(c))
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = f'A1:H{max(1, len(rows) + 1)}'
        widths = [18, 18, 30, 12, 14, 12, 12, 12]
        for i, w in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = w

    write_table(ws2, subscribers)
    write_table(ws3, overdue)

    out = io.BytesIO()
    wb.save(out)
    data = out.getvalue()
    resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="financial_report_{now.strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return resp


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_download_invoice(request, user_id: str):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    from django.http import HttpResponse

    u = User.objects.filter(id=user_id, is_superuser=False).select_related('subscription_onboarding').first()
    if not u:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    now = timezone.now()
    onboarding = getattr(u, 'subscription_onboarding', None)
    has_sub = bool(getattr(u, 'has_used_subscription', False))
    plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
    start_at = _payments_subscription_start(u, onboarding) if has_sub else None
    due_at = _payments_effective_due_at(u, start_at, plan)
    is_overdue = bool(has_sub and due_at and due_at < now)

    issued = now.date().isoformat()
    invoice_number = f'INV-{now.strftime("%Y%m%d")}-{u.id[-6:]}'
    amount_cents = plan['amount_cents'] if has_sub else 0

    lines = [
        'Daily Spanish',
        'Invoice',
        '',
        f'Invoice #: {invoice_number}',
        f'Issued on: {issued}',
        '',
        f'Bill to: {(u.name or "").strip() or "User"}',
        f'Email: {u.email}',
        '',
        f'Plan: {plan["plan_label"] if has_sub else "N/A"}',
        f'Amount: ${amount_cents / 100:.2f} USD',
        f'Status: {"Overdue" if is_overdue else "Paid" if has_sub else "No subscription record"}',
        f'Due date: {due_at.date().isoformat() if due_at else "N/A"}',
    ]
    pdf_bytes = _build_simple_pdf(lines)

    safe_email = (u.email or u.id).replace('@', '_at_').replace('/', '_')
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{safe_email}_{invoice_number}.pdf"'
    return resp


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_send_payment_reminder(request, user_id: str):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    from notifications.models import Notification, NotificationRecipient

    u = User.objects.filter(id=user_id, is_superuser=False).first()
    if not u:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    title = (request.data or {}).get('title') if isinstance(request.data, dict) else None
    message = (request.data or {}).get('message') if isinstance(request.data, dict) else None
    title = (title or 'Payment reminder').strip()
    message = (message or 'Your subscription payment is overdue. Please update your payment method to continue uninterrupted access.').strip()

    now = timezone.now()
    notif = Notification.objects.create(type='alert', title=title, message=message, audience_filters={'mode': 'user', 'user_id': u.id}, created_by=request.user)
    NotificationRecipient.objects.get_or_create(notification=notif, user=u, defaults={'delivered_at': now})
    notif.sent_at = now
    notif.save(update_fields=['sent_at'])
    return Response({'sent_to': 1}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_extend_payment_due(request, user_id: str):
    if not request.user.is_staff:
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    u = User.objects.filter(id=user_id, is_superuser=False).select_related('subscription_onboarding').first()
    if not u:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    data = request.data if isinstance(request.data, dict) else {}
    days_raw = data.get('days')
    due_date_raw = data.get('due_date')

    days = 7
    if days_raw not in [None, '']:
        try:
            days = int(days_raw)
        except Exception:
            return Response({'error': 'Invalid days'}, status=status.HTTP_400_BAD_REQUEST)
    if days < 1:
        return Response({'error': 'Days must be >= 1'}, status=status.HTTP_400_BAD_REQUEST)
    if days > 365:
        return Response({'error': 'Days too large'}, status=status.HTTP_400_BAD_REQUEST)

    now = timezone.now()
    onboarding = getattr(u, 'subscription_onboarding', None)
    plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))
    has_sub = bool(getattr(u, 'has_used_subscription', False))
    start_at = _payments_subscription_start(u, onboarding) if has_sub else None
    computed_due = _payments_effective_due_at(u, start_at, plan)

    if due_date_raw:
        try:
            from datetime import datetime as dt_datetime, timezone as dt_timezone

            target = dt_datetime.fromisoformat(str(due_date_raw))
            if timezone.is_naive(target):
                target = timezone.make_aware(target, timezone=dt_timezone.utc)
        except Exception:
            return Response({'error': 'Invalid due_date'}, status=status.HTTP_400_BAD_REQUEST)
        u.payment_due_override = target
        u.save(update_fields=['payment_due_override', 'updated_at'])
        return Response({'due_at': u.payment_due_override.isoformat()}, status=status.HTTP_200_OK)

    base = computed_due if computed_due else now
    if base < now:
        base = now
    u.payment_due_override = base + timedelta(days=days)
    u.save(update_fields=['payment_due_override', 'updated_at'])
    return Response({'due_at': u.payment_due_override.isoformat()}, status=status.HTTP_200_OK)


def _build_user_receipts(user: User):
    now = timezone.now()
    onboarding = getattr(user, 'subscription_onboarding', None)
    has_sub = bool(getattr(user, 'has_used_subscription', False))
    plan = _payments_plan_details(getattr(onboarding, 'plan_key', None))

    receipts = []

    if has_sub:
        issued_at = _payments_subscription_start(user, onboarding)
        invoice_number = f'INV-{issued_at.strftime("%Y")}-{user.id[-6:]}'
        receipt_id = f'receipt_{invoice_number}'
        course = f'Daily Spanish {plan["plan_label"]} Subscription'
        receipts.append(
            {
                'id': receipt_id,
                'date': issued_at.isoformat(),
                'course': course,
                'amount': f'${plan["amount_cents"] / 100:.2f}',
                'status': 'Paid',
                'invoice_id': invoice_number,
            }
        )

    if not has_sub and onboarding is not None:
        issued_at = onboarding.created_at or now
        invoice_number = f'INV-{issued_at.strftime("%Y")}-{user.id[-6:]}-P'
        receipt_id = f'receipt_{invoice_number}'
        course = f'Daily Spanish {plan["plan_label"]} Subscription'
        receipts.append(
            {
                'id': receipt_id,
                'date': issued_at.isoformat(),
                'course': course,
                'amount': f'${plan["amount_cents"] / 100:.2f}',
                'status': 'Pending',
                'invoice_id': invoice_number,
            }
        )

    return receipts


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_receipts(request):
    user = request.user
    return Response({'receipts': _build_user_receipts(user)}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_my_receipt(request, receipt_id: str):
    user = request.user
    receipts = _build_user_receipts(user)
    match = None
    for r in receipts:
        if r.get('id') == receipt_id:
            match = r
            break
    if match is None:
        return Response({'error': 'Receipt not found'}, status=status.HTTP_404_NOT_FOUND)

    from django.http import HttpResponse

    now = timezone.now()
    invoice_number = match.get('invoice_id') or f'INV-{now.strftime("%Y%m%d")}-{user.id[-6:]}'
    issued_on = match.get('date') or now.isoformat()
    course = match.get('course') or 'Daily Spanish Subscription'
    amount_raw = match.get('amount') or '$0.00'
    status = match.get('status') or 'Paid'

    lines = [
        'Daily Spanish',
        'Receipt',
        '',
        f'Invoice #: {invoice_number}',
        f'Issued on: {issued_on}',
        '',
        f'Billed to: {(user.name or "").strip() or "User"}',
        f'Email: {user.email}',
        '',
        f'Course/Package: {course}',
        f'Amount: {amount_raw} USD',
        f'Status: {status}',
    ]
    pdf_bytes = _build_simple_pdf(lines)

    safe_email = (user.email or user.id).replace('@', '_at_').replace('/', '_')
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{safe_email}_{invoice_number}.pdf"'
    return resp
